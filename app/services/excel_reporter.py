"""
exporter.py — Generador del Excel de resultados con openpyxl.

Produce un archivo .xlsx con tres pestañas:
  1. Líneas_OK       — líneas que cuadran perfectamente
  2. Incidencias     — líneas con discrepancias, coloreadas por severidad
  3. Resumen         — metadatos de la ejecución

Modo lote: añade la hoja Documentos y columnas de contexto (pedido/proveedor/archivos)
en Líneas_OK e Incidencias; Resumen con KPI del lote y tabla por par.
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .normalizer import ResultadoComparacion, MOTIVO_DESCRIPCION


@dataclass
class EntradaLote:
    """Un par pedido/confirmación dentro de un lote."""

    pedido_path: str
    confirmacion_path: str
    resultado: Optional[ResultadoComparacion] = None
    error: Optional[str] = None
    seconds: float = 0.0


# Colores
COLOR_HEADER_BG = "1F4E79"       # Azul oscuro
COLOR_HEADER_FG = "FFFFFF"       # Blanco
COLOR_OK_ROW = "E2EFDA"          # Verde claro
COLOR_ALTA = "FF0000"            # Rojo (MA-01, MA-02, MA-03, MA-06, MA-08, MA-09, MA-10)
COLOR_MEDIA = "FFC000"           # Naranja (MA-04, MA-05, MA-07, MA-11, MA-12)
COLOR_ALT_ROW = "F2F2F2"         # Gris alternado

INCIDENCIAS_ALTA = {"MA-01", "MA-02", "MA-03", "MA-06", "MA-08", "MA-09", "MA-10"}
INCIDENCIAS_MEDIA = {"MA-04", "MA-05", "MA-07", "MA-11", "MA-12"}
EXCEL_LOTE_SCHEMA_VERSION = "v1"


def exportar_excel(resultado: ResultadoComparacion, output_path: str) -> str:
    """
    Genera el Excel de resultados y lo guarda en output_path.
    Devuelve la ruta del archivo generado.
    """
    wb = Workbook()

    _sheet_lineas_ok(wb, resultado)
    _sheet_incidencias(wb, resultado)
    _sheet_resumen(wb, resultado)

    # Eliminar la hoja vacía por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path


def exportar_excel_extraccion(
    *,
    output_path: str,
    job_id: str,
    module: str,
    source_filename: str,
    data: dict,
    warnings: list[str],
) -> str:
    """
    Reporte simple para endpoint /extract (MVP).
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"

    summary_rows = [
        ("job_id", job_id),
        ("module", module),
        ("source_filename", source_filename),
        ("document_type", data.get("tipo")),
        ("supplier", (data.get("cabecera") or {}).get("proveedor_nombre")),
        ("lines_count", len(data.get("lineas", []))),
        ("warnings", " | ".join(warnings) if warnings else ""),
        ("generated_at_utc", datetime.utcnow().isoformat()),
    ]
    for idx, (k, v) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws_summary.cell(row=idx, column=2, value=v)
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 80

    ws_lines = wb.create_sheet("Lineas")
    lineas = data.get("lineas", []) or []
    columns = _infer_columns(lineas)
    if not columns:
        columns = ["info"]
        ws_lines.append(columns)
        ws_lines.append(["Sin líneas extraídas"])
    else:
        ws_lines.append(columns)
        for cidx, cname in enumerate(columns, start=1):
            cell = ws_lines.cell(row=1, column=cidx)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.font = Font(bold=True, color=COLOR_HEADER_FG)
            cell.alignment = Alignment(horizontal="center")
        for row in lineas:
            ws_lines.append([row.get(c) for c in columns])

    for idx, col in enumerate(columns, start=1):
        ws_lines.column_dimensions[get_column_letter(idx)].width = min(48, max(14, len(str(col)) + 4))

    wb.save(output_path)
    return output_path


def _infer_columns(lineas: list) -> list[str]:
    keys: list[str] = []
    seen = set()
    for row in lineas:
        if not isinstance(row, dict):
            continue
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                keys.append(str(k))
    return keys


def exportar_excel_lote(
    batch_id: str,
    started_at: str,
    finished_at: str,
    entradas: list[EntradaLote],
    output_path: str,
) -> str:
    """
    Un único workbook con hoja Documentos, líneas consolidadas y resumen de lote.
    """
    wb = Workbook()
    _sheet_documentos(wb, batch_id, started_at, finished_at, entradas)
    _sheet_lineas_ok_lote(wb, entradas)
    _sheet_incidencias_lote(wb, entradas)
    _sheet_resumen_lote(wb, batch_id, started_at, finished_at, entradas)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _assert_lote_schema_contract(wb)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Modo lote: Documentos + hojas con contexto
# ---------------------------------------------------------------------------

COLS_DOC = [
    ("#", 5),
    ("Archivo_pedido", 26),
    ("Archivo_confirmación", 26),
    ("Estado_carga", 12),
    ("Error", 48),
    ("Pedido_ref", 14),
    ("Proveedor", 30),
    ("Estado_comparación", 20),
    ("Líneas_OK", 11),
    ("Líneas_incidencia", 15),
    ("Segundos", 10),
]


def _sheet_documentos(
    wb: Workbook,
    batch_id: str,
    started_at: str,
    finished_at: str,
    entradas: list[EntradaLote],
):
    ws = wb.create_sheet("Documentos")
    ws.cell(row=1, column=1, value="batch_id").font = Font(bold=True)
    ws.cell(row=1, column=2, value=batch_id)
    ws.cell(row=2, column=1, value="inicio_UTC").font = Font(bold=True)
    ws.cell(row=2, column=2, value=started_at)
    ws.cell(row=3, column=1, value="fin_UTC").font = Font(bold=True)
    ws.cell(row=3, column=2, value=finished_at)
    start_row = 5
    _write_header_at_row(ws, COLS_DOC, row=start_row)

    for i, e in enumerate(entradas, start=1):
        r = start_row + i
        ok_load = e.error is None and e.resultado is not None
        estado_carga = "OK" if ok_load else "ERROR"
        ref = e.resultado.pedido_ref if e.resultado else ""
        prov = e.resultado.proveedor_nombre if e.resultado else ""
        est_comp = ""
        n_ok = ""
        n_inc = ""
        if e.resultado:
            est_comp = e.resultado.estado_global
            n_ok = len(e.resultado.lineas_ok)
            n_inc = len(e.resultado.lineas_incidencia)
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=Path(e.pedido_path).name)
        ws.cell(row=r, column=3, value=Path(e.confirmacion_path).name)
        ws.cell(row=r, column=4, value=estado_carga)
        ws.cell(row=r, column=5, value=e.error or "")
        ws.cell(row=r, column=6, value=ref)
        ws.cell(row=r, column=7, value=prov)
        ws.cell(row=r, column=8, value=est_comp)
        ws.cell(row=r, column=9, value=n_ok)
        ws.cell(row=r, column=10, value=n_inc)
        ws.cell(row=r, column=11, value=round(e.seconds, 2))

    _auto_width_from_row(ws, COLS_DOC, start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _write_header_at_row(ws, cols: list, row: int = 1):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FG)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, bottom=thin)
    for col_idx, (col_name, _) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 35


def _auto_width_from_row(ws, cols: list, header_row: int):
    for col_idx, (_, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Hoja 1: Líneas_OK (un solo par)
# ---------------------------------------------------------------------------

COLS_OK = [
    ("Cód. Proveedor", 18),
    ("Descripción Pedido", 40),
    ("Descripción Confirmación", 40),
    ("Cant. Pedido", 14),
    ("Cant. Confirm.", 14),
    ("Precio Pedido", 15),
    ("Precio Confirm.", 15),
    ("Dto. Pedido %", 14),
    ("Dto. Confirm. %", 15),
]


def _sheet_lineas_ok(wb: Workbook, resultado: ResultadoComparacion):
    ws = wb.create_sheet("Líneas_OK")
    _write_header(ws, COLS_OK)

    for i, l in enumerate(resultado.lineas_ok, start=2):
        row = [
            l.cod_proveedor,
            l.descripcion_pedido,
            l.descripcion_confirmacion or "",
            l.cantidad_pedido,
            l.cantidad_confirmacion or "",
            l.precio_pedido,
            l.precio_confirmacion or "",
            l.dto_pedido,
            l.dto_confirmacion if l.dto_confirmacion is not None else "",
        ]
        ws.append(row)
        fill = PatternFill("solid", fgColor=COLOR_OK_ROW if i % 2 == 0 else "FFFFFF")
        for cell in ws[i]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    _auto_width(ws, COLS_OK)


# ---------------------------------------------------------------------------
# Hoja 2: Incidencias
# ---------------------------------------------------------------------------

COLS_INC = [
    ("Cód. Proveedor", 18),
    ("Motivo(s)", 35),
    ("Descripción Motivo", 45),
    ("Descripción Pedido", 38),
    ("Descripción Confirmación", 38),
    ("Cant. Pedido", 13),
    ("Cant. Confirm.", 13),
    ("Precio Pedido", 14),
    ("Precio Confirm.", 14),
    ("Dto. Pedido %", 13),
    ("Dto. Confirm. %", 14),
]


def _sheet_incidencias(wb: Workbook, resultado: ResultadoComparacion):
    ws = wb.create_sheet("Incidencias")
    _write_header(ws, COLS_INC)

    for i, l in enumerate(resultado.lineas_incidencia, start=2):
        motivos_str = " | ".join(l.motivos)
        desc_motivos = " | ".join(MOTIVO_DESCRIPCION.get(m, m) for m in l.motivos)

        row = [
            l.cod_proveedor,
            motivos_str,
            desc_motivos,
            l.descripcion_pedido,
            l.descripcion_confirmacion or "",
            l.cantidad_pedido if l.cantidad_pedido else "",
            l.cantidad_confirmacion if l.cantidad_confirmacion is not None else "",
            l.precio_pedido if l.precio_pedido else "",
            l.precio_confirmacion if l.precio_confirmacion is not None else "",
            l.dto_pedido if l.dto_pedido is not None else "",
            l.dto_confirmacion if l.dto_confirmacion is not None else "",
        ]
        ws.append(row)

        # Color de fila según severidad máxima
        max_sev = _max_severidad(l.motivos)
        if max_sev == "alta":
            bg = COLOR_ALTA
            font_color = "FFFFFF"
        elif max_sev == "media":
            bg = COLOR_MEDIA
            font_color = "000000"
        else:
            bg = COLOR_ALT_ROW
            font_color = "000000"

        fill = PatternFill("solid", fgColor=bg)
        font = Font(color=font_color)
        for cell in ws[i]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    _auto_width(ws, COLS_INC)


COLS_CONTEXT = [
    ("Pedido_ref", 14),
    ("Proveedor", 28),
    ("Archivo_pedido", 26),
    ("Archivo_confirmación", 26),
]

COLS_OK_LOTE = COLS_CONTEXT + COLS_OK
COLS_INC_LOTE = COLS_CONTEXT + COLS_INC

EXPECTED_LOTE_SHEETS = ["Documentos", "Líneas_OK", "Incidencias", "Resumen"]
EXPECTED_RESUMEN_DETALLE_HEADERS = [
    "Archivo_pedido",
    "Archivo_confirmación",
    "Pedido_ref",
    "Proveedor",
    "Estado_carga",
    "Estado_comparación",
    "Error",
    "Segundos",
    "Líneas_OK",
    "Líneas_incidencia",
]


def _sheet_lineas_ok_lote(wb: Workbook, entradas: list[EntradaLote]):
    ws = wb.create_sheet("Líneas_OK")
    _write_header(ws, COLS_OK_LOTE)
    row_num = 2
    stripe = 0
    for e in entradas:
        if not e.resultado:
            continue
        pref = e.resultado.pedido_ref
        pnom = e.resultado.proveedor_nombre
        pfn = Path(e.pedido_path).name
        cfn = Path(e.confirmacion_path).name
        for l in e.resultado.lineas_ok:
            row = [
                pref,
                pnom,
                pfn,
                cfn,
                l.cod_proveedor,
                l.descripcion_pedido,
                l.descripcion_confirmacion or "",
                l.cantidad_pedido,
                l.cantidad_confirmacion or "",
                l.precio_pedido,
                l.precio_confirmacion or "",
                l.dto_pedido,
                l.dto_confirmacion if l.dto_confirmacion is not None else "",
            ]
            ws.append(row)
            fill = PatternFill(
                "solid", fgColor=COLOR_OK_ROW if stripe % 2 == 0 else "FFFFFF"
            )
            for cell in ws[row_num]:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
            row_num += 1
            stripe += 1
    _auto_width(ws, COLS_OK_LOTE)


def _sheet_incidencias_lote(wb: Workbook, entradas: list[EntradaLote]):
    ws = wb.create_sheet("Incidencias")
    _write_header(ws, COLS_INC_LOTE)
    row_num = 2
    for e in entradas:
        if not e.resultado:
            continue
        pref = e.resultado.pedido_ref
        pnom = e.resultado.proveedor_nombre
        pfn = Path(e.pedido_path).name
        cfn = Path(e.confirmacion_path).name
        for l in e.resultado.lineas_incidencia:
            motivos_str = " | ".join(l.motivos)
            desc_motivos = " | ".join(MOTIVO_DESCRIPCION.get(m, m) for m in l.motivos)
            row = [
                pref,
                pnom,
                pfn,
                cfn,
                l.cod_proveedor,
                motivos_str,
                desc_motivos,
                l.descripcion_pedido,
                l.descripcion_confirmacion or "",
                l.cantidad_pedido if l.cantidad_pedido else "",
                l.cantidad_confirmacion if l.cantidad_confirmacion is not None else "",
                l.precio_pedido if l.precio_pedido else "",
                l.precio_confirmacion if l.precio_confirmacion is not None else "",
                l.dto_pedido if l.dto_pedido is not None else "",
                l.dto_confirmacion if l.dto_confirmacion is not None else "",
            ]
            ws.append(row)
            max_sev = _max_severidad(l.motivos)
            if max_sev == "alta":
                bg = COLOR_ALTA
                font_color = "FFFFFF"
            elif max_sev == "media":
                bg = COLOR_MEDIA
                font_color = "000000"
            else:
                bg = COLOR_ALT_ROW
                font_color = "000000"
            fill = PatternFill("solid", fgColor=bg)
            font = Font(color=font_color)
            for cell in ws[row_num]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            row_num += 1
    ws.row_dimensions[1].height = 30
    _auto_width(ws, COLS_INC_LOTE)


def _sheet_resumen_lote(
    wb: Workbook,
    batch_id: str,
    started_at: str,
    finished_at: str,
    entradas: list[EntradaLote],
):
    ws = wb.create_sheet("Resumen")
    total_pairs = len(entradas)
    pairs_error = sum(1 for e in entradas if e.error is not None or e.resultado is None)
    pairs_ok_parse = total_pairs - pairs_error
    pairs_inc = sum(
        1
        for e in entradas
        if e.resultado and e.resultado.estado_global == "CON INCIDENCIAS"
    )
    lines_ok = sum(len(e.resultado.lineas_ok) for e in entradas if e.resultado)
    lines_inc = sum(len(e.resultado.lineas_incidencia) for e in entradas if e.resultado)

    datos = [
        ("Lote (batch_id)", batch_id),
        ("Inicio (UTC)", started_at),
        ("Fin (UTC)", finished_at),
        ("", ""),
        ("Pares en lote", total_pairs),
        ("Pares con error de carga/parseo", pairs_error),
        ("Pares comparados", pairs_ok_parse),
        ("Pares con incidencias en datos", pairs_inc),
        ("Total líneas OK (todas las comparaciones)", lines_ok),
        ("Total líneas con incidencia", lines_inc),
    ]

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FG)
    label_font = Font(bold=True)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 28

    for r_idx, (label, value) in enumerate(datos, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=label)
        cell_b = ws.cell(row=r_idx, column=2, value=value)
        if r_idx == 1:
            cell_a.fill = header_fill
            cell_a.font = header_font
            cell_b.fill = header_fill
            cell_b.font = header_font
        elif label and label != "":
            cell_a.font = label_font
        cell_a.alignment = Alignment(vertical="center")
        cell_b.alignment = Alignment(vertical="center")

    detalle_row = len(datos) + 2
    ws.cell(row=detalle_row, column=1, value="Detalle por par").font = Font(bold=True, size=12)
    hdr = detalle_row + 1
    cols_det = [
        "Archivo_pedido",
        "Archivo_confirmación",
        "Pedido_ref",
        "Proveedor",
        "Estado_carga",
        "Estado_comparación",
        "Error",
        "Segundos",
        "Líneas_OK",
        "Líneas_incidencia",
    ]
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, bottom=thin, top=thin)
    for c, name in enumerate(cols_det, start=1):
        cell = ws.cell(row=hdr, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 28

    for i, e in enumerate(entradas, start=1):
        r = hdr + i
        ok_load = e.error is None and e.resultado is not None
        ws.cell(row=r, column=1, value=Path(e.pedido_path).name)
        ws.cell(row=r, column=2, value=Path(e.confirmacion_path).name)
        ws.cell(row=r, column=3, value=e.resultado.pedido_ref if e.resultado else "")
        ws.cell(row=r, column=4, value=e.resultado.proveedor_nombre if e.resultado else "")
        ws.cell(row=r, column=5, value="OK" if ok_load else "ERROR")
        ws.cell(
            row=r,
            column=6,
            value=e.resultado.estado_global if e.resultado else "",
        )
        ws.cell(row=r, column=7, value=e.error or "")
        ws.cell(row=r, column=8, value=round(e.seconds, 2))
        n_ok = len(e.resultado.lineas_ok) if e.resultado else ""
        n_inc = len(e.resultado.lineas_incidencia) if e.resultado else ""
        ws.cell(row=r, column=9, value=n_ok)
        ws.cell(row=r, column=10, value=n_inc)

    for c in range(1, len(cols_det) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["G"].width = 40


# ---------------------------------------------------------------------------
# Hoja 3: Resumen (un solo par)
# ---------------------------------------------------------------------------

def _sheet_resumen(wb: Workbook, resultado: ResultadoComparacion):
    ws = wb.create_sheet("Resumen")

    datos = [
        ("Pedido Aquacenter", resultado.pedido_ref),
        ("Proveedor", resultado.proveedor_nombre),
        ("Estado global", resultado.estado_global),
        ("Fecha ejecución", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Líneas en pedido", resultado.total_lineas_pedido),
        ("Líneas en confirmación", resultado.total_lineas_confirmacion),
        ("Líneas procesadas", resultado.total_lineas),
        ("Líneas OK", len(resultado.lineas_ok)),
        ("Líneas con incidencia", len(resultado.lineas_incidencia)),
        ("", ""),
        ("Desglose de incidencias", ""),
    ]

    # Contar por código de motivo
    conteo: dict[str, int] = {}
    for l in resultado.lineas_incidencia:
        for m in l.motivos:
            conteo[m] = conteo.get(m, 0) + 1

    for cod, n in sorted(conteo.items()):
        datos.append((f"  {cod} — {MOTIVO_DESCRIPCION.get(cod, cod)}", n))

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FG)
    label_font = Font(bold=True)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30

    for r_idx, (label, value) in enumerate(datos, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=label)
        cell_b = ws.cell(row=r_idx, column=2, value=value)

        if r_idx == 1:
            cell_a.fill = header_fill
            cell_a.font = header_font
            cell_b.fill = header_fill
            cell_b.font = header_font
        elif label and not label.startswith(" ") and label != "":
            cell_a.font = label_font

        # Destacar estado global
        if label == "Estado global":
            color = "FF0000" if value == "CON INCIDENCIAS" else "375623"
            cell_b.font = Font(bold=True, color=color)

        cell_a.alignment = Alignment(vertical="center")
        cell_b.alignment = Alignment(vertical="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_header(ws, cols: list):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FG)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, bottom=thin)

    for col_idx, (col_name, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"


def _auto_width(ws, cols: list):
    for col_idx, (_, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _max_severidad(motivos: list) -> str:
    for m in motivos:
        if m in INCIDENCIAS_ALTA:
            return "alta"
    for m in motivos:
        if m in INCIDENCIAS_MEDIA:
            return "media"
    return "baja"


def _assert_lote_schema_contract(wb: Workbook) -> None:
    """
    Contrato estable para automatizaciones (Make/n8n/Sheets).
    Si alguien altera hojas/columnas por accidente, falla explícitamente.
    """
    if wb.sheetnames != EXPECTED_LOTE_SHEETS:
        raise ValueError(
            "Contrato Excel lote roto: hojas esperadas "
            f"{EXPECTED_LOTE_SHEETS}, obtenido {wb.sheetnames} "
            f"(schema {EXCEL_LOTE_SCHEMA_VERSION})."
        )

    ws_doc = wb["Documentos"]
    ws_ok = wb["Líneas_OK"]
    ws_inc = wb["Incidencias"]
    ws_res = wb["Resumen"]

    headers_doc = [ws_doc.cell(row=5, column=i).value for i in range(1, len(COLS_DOC) + 1)]
    expected_doc = [c for c, _ in COLS_DOC]
    if headers_doc != expected_doc:
        raise ValueError(
            "Contrato Excel lote roto en 'Documentos' "
            f"(schema {EXCEL_LOTE_SCHEMA_VERSION})."
        )

    headers_ok = [ws_ok.cell(row=1, column=i).value for i in range(1, len(COLS_OK_LOTE) + 1)]
    expected_ok = [c for c, _ in COLS_OK_LOTE]
    if headers_ok != expected_ok:
        raise ValueError(
            "Contrato Excel lote roto en 'Líneas_OK' "
            f"(schema {EXCEL_LOTE_SCHEMA_VERSION})."
        )

    headers_inc = [ws_inc.cell(row=1, column=i).value for i in range(1, len(COLS_INC_LOTE) + 1)]
    expected_inc = [c for c, _ in COLS_INC_LOTE]
    if headers_inc != expected_inc:
        raise ValueError(
            "Contrato Excel lote roto en 'Incidencias' "
            f"(schema {EXCEL_LOTE_SCHEMA_VERSION})."
        )

    headers_res = [
        ws_res.cell(row=13, column=i).value
        for i in range(1, len(EXPECTED_RESUMEN_DETALLE_HEADERS) + 1)
    ]
    if headers_res != EXPECTED_RESUMEN_DETALLE_HEADERS:
        raise ValueError(
            "Contrato Excel lote roto en 'Resumen' (detalle por par) "
            f"(schema {EXCEL_LOTE_SCHEMA_VERSION})."
        )
