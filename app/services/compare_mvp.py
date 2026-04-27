from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.supplier_cleaner import clean_supplier_name

QTY_TOLERANCE = 0.001
UNIT_PRICE_TOLERANCE = 0.02
DISCOUNT_TOLERANCE = 0.01
LINE_TOTAL_TOLERANCE = 0.02


def compare_documents_mvp(
    *,
    origin_doc: dict[str, Any],
    target_doc: dict[str, Any],
    job_id: str,
    module: str,
    output_dir: str,
) -> dict[str, Any]:
    origin_type = origin_doc.get("tipo", "desconocido")
    target_type = target_doc.get("tipo", "desconocido")
    supplier = clean_supplier_name(
        (origin_doc.get("cabecera") or {}).get("proveedor_nombre")
        or (target_doc.get("cabecera") or {}).get("proveedor_nombre")
    )

    origin_lines = [_normalize_line(line, side="origin") for line in (origin_doc.get("lineas") or [])]
    target_lines = [_normalize_line(line, side="target") for line in (target_doc.get("lineas") or [])]

    incidents: list[dict[str, Any]] = []
    ok_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []

    origin_by_code: dict[str, list[dict[str, Any]]] = {}
    target_by_code: dict[str, list[dict[str, Any]]] = {}

    for line in origin_lines:
        code = line["supplier_code"]
        if not code:
            unmatched_rows.append(_unmatched_row("origin", "MISSING_SUPPLIER_CODE", line))
            incidents.append(_incident("MISSING_SUPPLIER_CODE", line, None, "Línea de origen sin supplier_code"))
            continue
        origin_by_code.setdefault(code, []).append(line)

    for line in target_lines:
        code = line["supplier_code"]
        if not code:
            unmatched_rows.append(_unmatched_row("target", "MISSING_SUPPLIER_CODE", line))
            incidents.append(_incident("MISSING_SUPPLIER_CODE", None, line, "Línea de destino sin supplier_code"))
            continue
        target_by_code.setdefault(code, []).append(line)

    all_codes = sorted(set(origin_by_code.keys()) | set(target_by_code.keys()))
    for code in all_codes:
        ol = origin_by_code.get(code, [])
        tl = target_by_code.get(code, [])

        pairs = min(len(ol), len(tl))
        for i in range(pairs):
            origin_line = ol[i]
            target_line = tl[i]

            mismatch_fields: list[dict[str, Any]] = []
            _compare_field(mismatch_fields, "quantity", origin_line["quantity"], target_line["quantity"], QTY_TOLERANCE)
            _compare_field(
                mismatch_fields,
                "unit_price",
                origin_line["unit_price"],
                target_line["unit_price"],
                UNIT_PRICE_TOLERANCE,
            )
            _compare_field(
                mismatch_fields,
                "discount",
                origin_line["discount"],
                target_line["discount"],
                DISCOUNT_TOLERANCE,
            )
            _compare_field(
                mismatch_fields,
                "line_total",
                origin_line["line_total"],
                target_line["line_total"],
                LINE_TOTAL_TOLERANCE,
            )

            if mismatch_fields:
                incidents.append(
                    _incident(
                        "VALUE_MISMATCH",
                        origin_line,
                        target_line,
                        "Diferencias en: " + ", ".join(x["field"] for x in mismatch_fields),
                        mismatch_fields,
                    )
                )
            else:
                ok_rows.append({"supplier_code": code, "origin": origin_line, "target": target_line})

        if len(ol) > pairs:
            for extra in ol[pairs:]:
                unmatched_rows.append(_unmatched_row("origin", "NO_MATCH_TARGET", extra))
                incidents.append(_incident("NO_MATCH_TARGET", extra, None, "Línea origen sin match en destino"))

        if len(tl) > pairs:
            for extra in tl[pairs:]:
                unmatched_rows.append(_unmatched_row("target", "NO_MATCH_ORIGIN", extra))
                incidents.append(_incident("NO_MATCH_ORIGIN", None, extra, "Línea destino sin match en origen"))

    output_file = f"{job_id}.xlsx"
    output_path = Path(output_dir) / output_file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _export_compare_excel(
        output_path=str(output_path),
        job_id=job_id,
        module=module,
        supplier=supplier,
        origin_document_type=origin_type,
        target_document_type=target_type,
        origin_count=len(origin_lines),
        target_count=len(target_lines),
        ok_rows=ok_rows,
        incidents=incidents,
        unmatched_rows=unmatched_rows,
    )

    return {
        "success": True,
        "job_id": job_id,
        "origin_document_type": origin_type,
        "target_document_type": target_type,
        "supplier": supplier,
        "lines_origin": len(origin_lines),
        "lines_target": len(target_lines),
        "lines_ok": len(ok_rows),
        "incidents_total": len(incidents),
        "incidents": incidents,
        "output_excel": f"/outputs/{output_file}",
    }


def _normalize_line(line: dict[str, Any], *, side: str) -> dict[str, Any]:
    code = str((line.get("cod_proveedor") or "")).strip().upper()
    quantity = _as_float(line.get("cantidad"))
    unit_price = _as_float(
        line.get("precio_unitario") if side == "target" else line.get("precio_vta", line.get("precio_unitario"))
    )
    discount = _as_float(line.get("dto")) or 0.0
    line_total = _as_float(line.get("importe"))
    if line_total is None and quantity is not None and unit_price is not None:
        line_total = quantity * unit_price * (1 - discount / 100.0)

    return {
        "supplier_code": code,
        "description": str(line.get("descripcion") or "").strip(),
        "quantity": quantity,
        "unit_price": unit_price,
        "discount": discount,
        "line_total": line_total,
    }


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _approx_equal(a: float | None, b: float | None, tol: float) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def _compare_field(mismatch_fields: list[dict[str, Any]], field: str, a: float | None, b: float | None, tol: float) -> None:
    if not _approx_equal(a, b, tol):
        mismatch_fields.append({"field": field, "origin": a, "target": b})


def _incident(
    code: str,
    origin_line: dict[str, Any] | None,
    target_line: dict[str, Any] | None,
    message: str,
    mismatch_fields: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    supplier_code = (origin_line or target_line or {}).get("supplier_code")
    return {
        "code": code,
        "supplier_code": supplier_code,
        "message": message,
        "mismatch_fields": mismatch_fields or [],
        "origin_line": origin_line,
        "target_line": target_line,
    }


def _unmatched_row(side: str, reason: str, line: dict[str, Any]) -> dict[str, Any]:
    return {"side": side, "reason": reason, "line": line}


def _export_compare_excel(
    *,
    output_path: str,
    job_id: str,
    module: str,
    supplier: str | None,
    origin_document_type: str,
    target_document_type: str,
    origin_count: int,
    target_count: int,
    ok_rows: list[dict[str, Any]],
    incidents: list[dict[str, Any]],
    unmatched_rows: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    ws_summary = wb.active
    ws_summary.title = "Resumen"
    summary_rows = [
        ("job_id", job_id),
        ("module", module),
        ("supplier", supplier or ""),
        ("origin_document_type", origin_document_type),
        ("target_document_type", target_document_type),
        ("lines_origin", origin_count),
        ("lines_target", target_count),
        ("lines_ok", len(ok_rows)),
        ("incidents_total", len(incidents)),
        ("generated_at_utc", datetime.utcnow().isoformat()),
    ]
    for i, (k, v) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=v)
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 80

    ws_ok = wb.create_sheet("Lineas_OK")
    ok_headers = [
        "supplier_code",
        "description_origin",
        "description_target",
        "quantity_origin",
        "quantity_target",
        "unit_price_origin",
        "unit_price_target",
        "discount_origin",
        "discount_target",
        "line_total_origin",
        "line_total_target",
    ]
    ws_ok.append(ok_headers)
    _style_header(ws_ok, header_fill, header_font)
    for row in ok_rows:
        o = row["origin"]
        t = row["target"]
        ws_ok.append(
            [
                row["supplier_code"],
                o.get("description"),
                t.get("description"),
                o.get("quantity"),
                t.get("quantity"),
                o.get("unit_price"),
                t.get("unit_price"),
                o.get("discount"),
                t.get("discount"),
                o.get("line_total"),
                t.get("line_total"),
            ]
        )

    ws_inc = wb.create_sheet("Incidencias")
    inc_headers = [
        "code",
        "supplier_code",
        "message",
        "mismatch_fields",
        "quantity_origin",
        "quantity_target",
        "unit_price_origin",
        "unit_price_target",
        "discount_origin",
        "discount_target",
        "line_total_origin",
        "line_total_target",
    ]
    ws_inc.append(inc_headers)
    _style_header(ws_inc, header_fill, header_font)
    for inc in incidents:
        o = inc.get("origin_line") or {}
        t = inc.get("target_line") or {}
        ws_inc.append(
            [
                inc.get("code"),
                inc.get("supplier_code"),
                inc.get("message"),
                ", ".join(x.get("field") for x in (inc.get("mismatch_fields") or [])),
                o.get("quantity"),
                t.get("quantity"),
                o.get("unit_price"),
                t.get("unit_price"),
                o.get("discount"),
                t.get("discount"),
                o.get("line_total"),
                t.get("line_total"),
            ]
        )

    ws_un = wb.create_sheet("No_Emparejadas")
    un_headers = [
        "side",
        "reason",
        "supplier_code",
        "description",
        "quantity",
        "unit_price",
        "discount",
        "line_total",
    ]
    ws_un.append(un_headers)
    _style_header(ws_un, header_fill, header_font)
    for row in unmatched_rows:
        line = row.get("line") or {}
        ws_un.append(
            [
                row.get("side"),
                row.get("reason"),
                line.get("supplier_code"),
                line.get("description"),
                line.get("quantity"),
                line.get("unit_price"),
                line.get("discount"),
                line.get("line_total"),
            ]
        )

    for ws in (ws_ok, ws_inc, ws_un):
        _auto_width(ws)

    wb.save(output_path)


def _style_header(ws, fill: PatternFill, font: Font) -> None:
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws) -> None:
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, max_len + 2))

