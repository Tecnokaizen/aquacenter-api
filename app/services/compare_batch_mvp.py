from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.compare_runtime import run_compare_pair
from app.services.pdf_extractor import detect_doc_type, extract_pdf_document
from app.services.supplier_cleaner import clean_supplier_name


@dataclass
class BatchInputDocument:
    filename: str
    path: str


@dataclass
class ClassifiedDocument:
    filename: str
    path: str
    doc_type: str
    supplier: str | None
    reference: str | None
    line_codes: set[str]
    warnings: list[str]


def run_compare_batch_mvp(
    *,
    batch_id: str,
    module: str,
    use_ai: bool,
    input_documents: list[BatchInputDocument],
    output_dir: str,
) -> dict[str, Any]:
    classified_docs: list[ClassifiedDocument] = []
    unmatched_documents: list[dict[str, Any]] = []

    for item in input_documents:
        doc_type = detect_doc_type(item.path)
        parsed_doc, warnings = extract_pdf_document(item.path, use_ai=use_ai)
        parsed_type = parsed_doc.get("tipo", doc_type)
        supplier = clean_supplier_name((parsed_doc.get("cabecera") or {}).get("proveedor_nombre"))
        reference = _extract_reference(parsed_doc)
        line_codes = _extract_line_codes(parsed_doc)

        classified = ClassifiedDocument(
            filename=item.filename,
            path=item.path,
            doc_type=parsed_type,
            supplier=supplier,
            reference=reference,
            line_codes=line_codes,
            warnings=warnings,
        )
        if parsed_type not in {"pedido", "confirmacion"}:
            unmatched_documents.append(
                {
                    "filename": item.filename,
                    "document_type": parsed_type,
                    "supplier": supplier,
                    "reason": "UNSUPPORTED_DOCUMENT_TYPE",
                    "warnings": warnings,
                }
            )
            continue
        classified_docs.append(classified)

    pedidos = [d for d in classified_docs if d.doc_type == "pedido"]
    confirmaciones = [d for d in classified_docs if d.doc_type == "confirmacion"]
    pairs, unmatched_indices = _pair_documents(pedidos, confirmaciones)

    for idx in sorted(unmatched_indices["pedidos"]):
        doc = pedidos[idx]
        unmatched_documents.append(
            {
                "filename": doc.filename,
                "document_type": doc.doc_type,
                "supplier": doc.supplier,
                "reason": "NO_MATCH_CONFIRMACION",
                "warnings": doc.warnings,
            }
        )
    for idx in sorted(unmatched_indices["confirmaciones"]):
        doc = confirmaciones[idx]
        unmatched_documents.append(
            {
                "filename": doc.filename,
                "document_type": doc.doc_type,
                "supplier": doc.supplier,
                "reason": "NO_MATCH_PEDIDO",
                "warnings": doc.warnings,
            }
        )

    pair_results: list[dict[str, Any]] = []
    comparisons_ok = 0
    comparisons_with_incidents = 0
    incidents_total = 0

    for i, (idx_pedido, idx_confirm, score) in enumerate(pairs, start=1):
        pedido = pedidos[idx_pedido]
        confirm = confirmaciones[idx_confirm]
        pair_job_id = f"{batch_id}_{i:03d}"
        payload, warnings = run_compare_pair(
            origin_path=pedido.path,
            target_path=confirm.path,
            job_id=pair_job_id,
            module=module,
            use_ai=use_ai,
            output_dir=output_dir,
        )
        if warnings:
            payload["warnings"] = warnings
        overall_status = payload.get("overall_status", "with_incidents")
        incidents_count = int(payload.get("incidents_total", 0) or 0)
        if overall_status == "ok":
            comparisons_ok += 1
        else:
            comparisons_with_incidents += 1
        incidents_total += incidents_count

        pair_results.append(
            {
                "pair_id": i,
                "score": score,
                "origin_file": pedido.filename,
                "target_file": confirm.filename,
                **payload,
            }
        )

    batch_excel_name = f"{batch_id}.xlsx"
    batch_excel_path = Path(output_dir) / batch_excel_name
    batch_excel_path.parent.mkdir(parents=True, exist_ok=True)
    _export_batch_excel(
        output_path=str(batch_excel_path),
        batch_id=batch_id,
        module=module,
        documents_total=len(input_documents),
        pair_results=pair_results,
        unmatched_documents=unmatched_documents,
        comparisons_ok=comparisons_ok,
        comparisons_with_incidents=comparisons_with_incidents,
        incidents_total=incidents_total,
    )

    return {
        "batch_id": batch_id,
        "documents_total": len(input_documents),
        "pairs_detected": len(pair_results),
        "comparisons_ok": comparisons_ok,
        "comparisons_with_incidents": comparisons_with_incidents,
        "unmatched_documents": unmatched_documents,
        "incidents_total": incidents_total,
        "batch_excel": f"/outputs/{batch_excel_name}",
        "pairs": pair_results,
    }


def _extract_reference(parsed_doc: dict[str, Any]) -> str | None:
    cabecera = parsed_doc.get("cabecera") or {}
    raw = cabecera.get("pedido") or cabecera.get("referencia_cliente") or cabecera.get("pedido_proveedor")
    if raw is None:
        return None
    normalized = "".join(ch for ch in str(raw).upper() if ch.isalnum())
    return normalized or None


def _extract_line_codes(parsed_doc: dict[str, Any]) -> set[str]:
    codes: set[str] = set()
    for line in parsed_doc.get("lineas") or []:
        code = str(line.get("cod_proveedor") or "").strip().upper()
        if code:
            codes.add(code)
    return codes


def _pair_documents(
    pedidos: list[ClassifiedDocument],
    confirmaciones: list[ClassifiedDocument],
) -> tuple[list[tuple[int, int, int]], dict[str, set[int]]]:
    candidates: list[tuple[int, int, int]] = []
    for i, pedido in enumerate(pedidos):
        for j, confirm in enumerate(confirmaciones):
            score = _pair_score(pedido, confirm)
            if score > 0:
                candidates.append((i, j, score))

    candidates.sort(key=lambda item: item[2], reverse=True)
    matched_pedidos: set[int] = set()
    matched_confirmaciones: set[int] = set()
    pairs: list[tuple[int, int, int]] = []

    for i, j, score in candidates:
        if i in matched_pedidos or j in matched_confirmaciones:
            continue
        matched_pedidos.add(i)
        matched_confirmaciones.add(j)
        pairs.append((i, j, score))

    unmatched = {
        "pedidos": set(range(len(pedidos))) - matched_pedidos,
        "confirmaciones": set(range(len(confirmaciones))) - matched_confirmaciones,
    }
    return pairs, unmatched


def _pair_score(pedido: ClassifiedDocument, confirm: ClassifiedDocument) -> int:
    score = 0
    if pedido.supplier and confirm.supplier and pedido.supplier == confirm.supplier:
        score += 40

    if pedido.reference and confirm.reference:
        if pedido.reference == confirm.reference:
            score += 60
        elif pedido.reference in confirm.reference or confirm.reference in pedido.reference:
            score += 30

    overlap = len(pedido.line_codes & confirm.line_codes)
    if overlap > 0:
        score += min(50, overlap * 10)

    return score


def _export_batch_excel(
    *,
    output_path: str,
    batch_id: str,
    module: str,
    documents_total: int,
    pair_results: list[dict[str, Any]],
    unmatched_documents: list[dict[str, Any]],
    comparisons_ok: int,
    comparisons_with_incidents: int,
    incidents_total: int,
) -> None:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    ws_summary = wb.active
    ws_summary.title = "Resumen"
    summary_rows = [
        ("batch_id", batch_id),
        ("module", module),
        ("documents_total", documents_total),
        ("pairs_detected", len(pair_results)),
        ("comparisons_ok", comparisons_ok),
        ("comparisons_with_incidents", comparisons_with_incidents),
        ("unmatched_documents", len(unmatched_documents)),
        ("incidents_total", incidents_total),
        ("generated_at_utc", datetime.utcnow().isoformat()),
    ]
    for idx, (k, v) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws_summary.cell(row=idx, column=2, value=v)
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 80

    ws_pairs = wb.create_sheet("Comparaciones")
    pair_headers = [
        "pair_id",
        "origin_file",
        "target_file",
        "supplier",
        "lines_origin",
        "lines_target",
        "lines_ok",
        "incidents_total",
        "overall_status",
        "output_excel",
    ]
    ws_pairs.append(pair_headers)
    _style_header(ws_pairs, header_fill, header_font)
    for row in pair_results:
        ws_pairs.append(
            [
                row.get("pair_id"),
                row.get("origin_file"),
                row.get("target_file"),
                row.get("supplier"),
                row.get("lines_origin"),
                row.get("lines_target"),
                row.get("lines_ok"),
                row.get("incidents_total"),
                row.get("overall_status"),
                row.get("output_excel"),
            ]
        )

    ws_incidents = wb.create_sheet("Incidencias")
    incident_headers = [
        "pair_id",
        "code",
        "supplier_code",
        "message",
        "mismatch_fields",
    ]
    ws_incidents.append(incident_headers)
    _style_header(ws_incidents, header_fill, header_font)
    for pair in pair_results:
        for inc in pair.get("incidents") or []:
            ws_incidents.append(
                [
                    pair.get("pair_id"),
                    inc.get("code"),
                    inc.get("supplier_code"),
                    inc.get("message"),
                    ", ".join(x.get("field") for x in (inc.get("mismatch_fields") or [])),
                ]
            )

    ws_unmatched = wb.create_sheet("No_Emparejados")
    unmatched_headers = [
        "filename",
        "document_type",
        "supplier",
        "reason",
        "warnings",
    ]
    ws_unmatched.append(unmatched_headers)
    _style_header(ws_unmatched, header_fill, header_font)
    for item in unmatched_documents:
        ws_unmatched.append(
            [
                item.get("filename"),
                item.get("document_type"),
                item.get("supplier"),
                item.get("reason"),
                ", ".join(item.get("warnings") or []),
            ]
        )

    for ws in (ws_pairs, ws_incidents, ws_unmatched):
        _auto_width(ws)

    wb.save(output_path)


def _style_header(ws, fill: PatternFill, font: Font) -> None:
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws) -> None:
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(i)].width = min(56, max(12, max_len + 2))

